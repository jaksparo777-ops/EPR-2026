import csv
import io
from django.http import HttpResponse
from django.db import transaction
from django.utils import timezone
from apps.products.models import Item, Client
from apps.workforce.models import Worker

# Column definitions
ITEM_HEADERS = [
    "Item Code*", "Item Name*", "Category", "Sub Category", "Material", "Variant",
    "Casting Weight (kg)", "Machining Weight (kg)", "Rate Per Piece", "Lot Size", "Lot With Box",
    "Casting Required (Y/N)", "Machining Required (Y/N)", "Polishing Required (Y/N)", "Packing Required (Y/N)", "Notes",
    "Client Name", "Companies",
    "Casting Worker Name", "Casting Rate",
    "Machining Worker Name", "Machining Rate",
    "Polishing Worker Name", "Polishing Rate",
    "Packing Worker Name", "Packing Rate",
    "Job Worker Name", "Job Worker Rate"
]

WORKER_HEADERS = [
    "Name*", "Employee ID", "Phone", "Designation", "Process (casting/machining/polishing/packaging)*",
    "Daily Rate", "Standard Shift Hours", "Salary Model (DAILY/FIXED)*", "Monthly Fixed Salary", "Monthly Allowance", "Overtime Rate",
    "Identity Number", "Emergency Contact Name", "Emergency Contact Phone", "Blood Group", "Rates (ItemCode:Rate,ItemCode:Rate,...)"
]

JOB_WORKER_HEADERS = [
    "Job Worker Code", "Name*", "Phone", "Email", "Address", "GST Number",
    "Process (casting/machining/polishing/packaging)*", "Rates (ItemCode:Rate,ItemCode:Rate,...)"
]


def generate_csv_template(headers, filename):
    """Generates a downloadable CSV template."""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
    writer = csv.writer(response)
    writer.writerow(headers)
    return response


def generate_xlsx_template(headers, filename):
    """Generates a downloadable Excel (.xlsx) template using openpyxl."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template"

    # Style definitions for header
    font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    alignment = Alignment(horizontal="center", vertical="center")

    ws.append(headers)

    # Apply styling to headers and adjust column widths
    for col_idx, col in enumerate(ws.columns, 1):
        cell = col[0]
        cell.font = font
        ws.row_dimensions[1].height = 25
        cell.fill = fill
        cell.alignment = alignment
        
        # Determine width based on header length
        header_len = len(str(cell.value or ''))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(header_len + 5, 15)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    return response


def parse_uploaded_file(file_obj):
    """
    Parses an uploaded CSV or XLSX file and returns headers and rows.
    """
    name = file_obj.name.lower()
    rows = []

    if name.endswith('.xlsx'):
        import openpyxl
        wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            # Check if it's completely empty row
            if all(val is None for val in row):
                continue
            rows.append([str(val).strip() if val is not None else "" for val in row])
    else:
        # Assume CSV
        try:
            content = file_obj.read().decode('utf-8-sig') # handle BOM marks cleanly
        except UnicodeDecodeError:
            content = file_obj.read().decode('latin-1')
            
        csv_file = io.StringIO(content)
        reader = csv.reader(csv_file)
        for row in reader:
            if not row or all(not cell.strip() for cell in row):
                continue
            rows.append([cell.strip() for cell in row])

    if not rows:
        return [], []
    
    headers = [h.strip() for h in rows[0]]
    return headers, rows[1:]


def to_float(val, default=0.0):
    if not val:
        return default
    try:
        return float(val)
    except ValueError:
        raise ValueError(f"'{val}' is not a valid decimal number.")


def to_int(val, default=0):
    if not val:
        return default
    try:
        return int(float(val)) # handles cases like "100.0" cleanly
    except ValueError:
        raise ValueError(f"'{val}' is not a valid whole number.")


def to_bool(val, default=True):
    if not val:
        return default
    val_clean = str(val).strip().upper()
    if val_clean in ('Y', 'YES', 'TRUE', '1', 'ON'):
        return True
    if val_clean in ('N', 'NO', 'FALSE', '0', 'OFF'):
        return False
    raise ValueError(f"'{val}' must be Y or N.")


def validate_items_data(rows):
    """
    Validates rows of Item data before import.
    Returns: list of dicts (parsed row + validation status)
    """
    validated_rows = []
    codes_in_file = set()

    for idx, row in enumerate(rows, 1):
        errors = []
        
        # Ensure row has enough columns
        if len(row) < 2:
            validated_rows.append({
                "row_idx": idx,
                "data": {"code": "N/A", "name": "N/A"},
                "action": "ERROR",
                "errors": ["Row has too few columns."]
            })
            continue

        # Extract columns safely
        code = row[0].strip()
        name = row[1].strip()
        category = row[2].strip() if len(row) > 2 else "OTHER"
        sub_category = row[3].strip() if len(row) > 3 else ""
        material = row[4].strip() if len(row) > 4 else ""
        variant = row[5].strip() if len(row) > 5 else ""
        
        raw_casting_w = row[6].strip() if len(row) > 6 else ""
        raw_machining_w = row[7].strip() if len(row) > 7 else ""
        raw_rate = row[8].strip() if len(row) > 8 else ""
        raw_lot_size = row[9].strip() if len(row) > 9 else ""
        raw_lot_box = row[10].strip() if len(row) > 10 else ""
        
        raw_cast_req = row[11].strip() if len(row) > 11 else "Y"
        raw_mach_req = row[12].strip() if len(row) > 12 else "Y"
        raw_polish_req = row[13].strip() if len(row) > 13 else "Y"
        raw_pack_req = row[14].strip() if len(row) > 14 else "Y"
        notes = row[15].strip() if len(row) > 15 else ""

        client_name = row[16].strip() if len(row) > 16 else ""
        companies_str = row[17].strip() if len(row) > 17 else ""
        
        casting_worker_name = row[18].strip() if len(row) > 18 else ""
        raw_casting_rate = row[19].strip() if len(row) > 19 else ""
        
        machining_worker_name = row[20].strip() if len(row) > 20 else ""
        raw_machining_rate = row[21].strip() if len(row) > 21 else ""
        
        polishing_worker_name = row[22].strip() if len(row) > 22 else ""
        raw_polishing_rate = row[23].strip() if len(row) > 23 else ""
        
        packing_worker_name = row[24].strip() if len(row) > 24 else ""
        raw_packing_rate = row[25].strip() if len(row) > 25 else ""
        
        job_worker_name = row[26].strip() if len(row) > 26 else ""
        raw_job_worker_rate = row[27].strip() if len(row) > 27 else ""

        # 1. Check Required Fields
        if not code:
            errors.append("Item Code is required.")
        if not name:
            errors.append("Item Name is required.")

        # 2. Check Deduplication within file
        if code:
            if code in codes_in_file:
                errors.append(f"Duplicate Item Code '{code}' in upload file.")
            else:
                codes_in_file.add(code)

        # 3. Numeric conversions & bounds checking
        casting_w = 0.0
        machining_w = 0.0
        rate = 0.0
        lot_size = 0
        lot_box = 0

        try:
            casting_w = to_float(raw_casting_w, 0.0)
            if casting_w < 0: errors.append("Casting weight cannot be negative.")
        except ValueError as e:
            errors.append(str(e))

        try:
            machining_w = to_float(raw_machining_w, 0.0)
            if machining_w < 0: errors.append("Machining weight cannot be negative.")
        except ValueError as e:
            errors.append(str(e))

        try:
            rate = to_float(raw_rate, 0.0)
            if rate < 0: errors.append("Rate per piece cannot be negative.")
        except ValueError as e:
            errors.append(str(e))

        try:
            lot_size = to_int(raw_lot_size, 0)
            if lot_size < 0: errors.append("Lot size cannot be negative.")
        except ValueError as e:
            errors.append(str(e))

        try:
            lot_box = to_int(raw_lot_box, 0)
            if lot_box < 0: errors.append("Lot with box cannot be negative.")
        except ValueError as e:
            errors.append(str(e))

        # 4. Boolean conversions
        cast_req = True
        mach_req = True
        polish_req = True
        pack_req = True

        try: cast_req = to_bool(raw_cast_req, True)
        except ValueError as e: errors.append(str(e))

        try: mach_req = to_bool(raw_mach_req, True)
        except ValueError as e: errors.append(str(e))

        try: polish_req = to_bool(raw_polish_req, True)
        except ValueError as e: errors.append(str(e))

        try: pack_req = to_bool(raw_pack_req, True)
        except ValueError as e: errors.append(str(e))

        # 5. Client & Companies lookup validation
        client_obj = None
        if client_name:
            client_obj = Client.objects.filter(name__iexact=client_name).first()
            if not client_obj:
                errors.append(f"Client '{client_name}' does not exist in the database.")
                
        company_ids = []
        if companies_str:
            from apps.client_orders.models import LegalEntity
            comp_names = [c.strip() for c in companies_str.split(',') if c.strip()]
            for cname in comp_names:
                comp_obj = LegalEntity.objects.filter(name__iexact=cname).first()
                if not comp_obj:
                    errors.append(f"Company '{cname}' does not exist in the database.")
                else:
                    company_ids.append(comp_obj.id)

        # 6. Rates & Workers lookup validation
        casting_rate = 0.0
        machining_rate = 0.0
        polishing_rate = 0.0
        packing_rate = 0.0
        job_worker_rate = 0.0
        
        try: casting_rate = to_float(raw_casting_rate, 0.0)
        except ValueError as e: errors.append(f"Casting Rate: {str(e)}")
        
        try: machining_rate = to_float(raw_machining_rate, 0.0)
        except ValueError as e: errors.append(f"Machining Rate: {str(e)}")
        
        try: polishing_rate = to_float(raw_polishing_rate, 0.0)
        except ValueError as e: errors.append(f"Polishing Rate: {str(e)}")
        
        try: packing_rate = to_float(raw_packing_rate, 0.0)
        except ValueError as e: errors.append(f"Packing Rate: {str(e)}")
        
        try: job_worker_rate = to_float(raw_job_worker_rate, 0.0)
        except ValueError as e: errors.append(f"Job Worker Rate: {str(e)}")

        casting_worker_id = None
        if casting_worker_name:
            cw = Worker.objects.filter(name__iexact=casting_worker_name).first()
            if not cw:
                errors.append(f"Casting Worker '{casting_worker_name}' not found.")
            else:
                casting_worker_id = cw.id
                
        machining_worker_id = None
        if machining_worker_name:
            mw = Worker.objects.filter(name__iexact=machining_worker_name).first()
            if not mw:
                errors.append(f"Machining Worker '{machining_worker_name}' not found.")
            else:
                machining_worker_id = mw.id
                
        polishing_worker_id = None
        if polishing_worker_name:
            pw = Worker.objects.filter(name__iexact=polishing_worker_name).first()
            if not pw:
                errors.append(f"Polishing Worker '{polishing_worker_name}' not found.")
            else:
                polishing_worker_id = pw.id
                
        packing_worker_id = None
        if packing_worker_name:
            pk = Worker.objects.filter(name__iexact=packing_worker_name).first()
            if not pk:
                errors.append(f"Packing Worker '{packing_worker_name}' not found.")
            else:
                packing_worker_id = pk.id
                
        job_worker_id = None
        if job_worker_name:
            from apps.workforce.models import JobWorker
            jw = JobWorker.objects.filter(name__iexact=job_worker_name).first()
            if not jw:
                errors.append(f"Job Worker '{job_worker_name}' not found.")
            else:
                job_worker_id = jw.id

        # 7. Check if Code exists in DB to determine UPDATE vs INSERT
        action = "INSERT"
        if code and not errors:
            exists = Item.objects.filter(code=code).exists()
            if exists:
                action = "UPDATE"

        parsed_data = {
            "code": code,
            "name": name,
            "category": category or "OTHER",
            "sub_category": sub_category,
            "material": material or "OTHER",
            "variant": variant,
            "casting_weight": casting_w,
            "machining_weight": machining_w,
            "rate_per_piece": rate,
            "lot_size": lot_size,
            "lot_with_box": lot_box,
            "casting_required": cast_req,
            "machining_required": mach_req,
            "polishing_required": polish_req,
            "packing_required": pack_req,
            "notes": notes,
            "client_id": client_obj.id if client_obj else None,
            "company_ids": company_ids,
            "casting_worker_id": casting_worker_id,
            "casting_rate": casting_rate,
            "machining_worker_id": machining_worker_id,
            "machining_rate": machining_rate,
            "polishing_worker_id": polishing_worker_id,
            "polishing_rate": polishing_rate,
            "packing_worker_id": packing_worker_id,
            "packing_rate": packing_rate,
            "job_worker_id": job_worker_id,
            "job_worker_rate": job_worker_rate
        }

        validated_rows.append({
            "row_idx": idx,
            "data": parsed_data,
            "action": "ERROR" if errors else action,
            "errors": errors
        })

    return validated_rows


def parse_piece_rates(raw_rates):
    """
    Parses raw rates string. Supports both:
    - Standard: "ITM001:10.5, ITM002:12.0"
    - Grouped: "ITM001,ITM002,ITM003:15.5"
    - Semicolon or comma separated: "ITM001,ITM002:10; ITM003:12"
    - Typo fallback: "3.K3" -> "3,K3"
    """
    import re
    # Replace dots followed by any non-digit with a comma (handles typo separators like '3.K3' -> '3,K3')
    raw_rates = re.sub(r'\.(\D)', r',\1', raw_rates)
    
    normalized = raw_rates.replace(';', ',').strip().rstrip(',')
    
    parts = normalized.split(':')
    if len(parts) < 2:
        return {}, [f"Invalid rate format: '{raw_rates}'. Must contain at least one ':'."]
        
    parsed_rates = {}
    errors = []
    
    # Split the first part by comma to get the starting list of item codes
    current_codes = [c.strip() for c in parts[0].split(',') if c.strip()]
    
    for i in range(1, len(parts)):
        sub_part = parts[i].strip()
        if i == len(parts) - 1:
            # Last part must be just the rate value
            rate_str = sub_part
            next_codes = []
        else:
            # Contains "Rate, NextItem1, NextItem2"
            sub_parts = [sp.strip() for sp in sub_part.split(',') if sp.strip()]
            if not sub_parts:
                errors.append("Missing rate value after colon.")
                continue
            rate_str = sub_parts[0]
            next_codes = sub_parts[1:]
            
        # Parse rate
        try:
            rate_float = float(rate_str)
            if rate_float < 0:
                errors.append(f"Rate value '{rate_str}' cannot be negative.")
            else:
                for code in current_codes:
                    if code:
                        parsed_rates[code] = rate_float
        except ValueError:
            errors.append(f"Invalid rate value '{rate_str}'.")
            
        current_codes = next_codes
        
    return parsed_rates, errors


def validate_workers_data(rows):
    """
    Validates rows of Worker data before import.
    Returns: list of dicts (parsed row + validation status)
    """
    validated_rows = []
    emp_ids_in_file = set()
    names_in_file = set()
    
    # Pre-fetch existing item codes from DB for validation
    item_codes = set(Item.objects.values_list('code', flat=True))

    for idx, row in enumerate(rows, 1):
        errors = []
        
        # Ensure row has enough columns
        if len(row) < 5:
            validated_rows.append({
                "row_idx": idx,
                "data": {"name": "N/A", "employee_id": "N/A"},
                "action": "ERROR",
                "errors": ["Row has too few columns."]
            })
            continue

        # Extract columns safely
        name = row[0].strip()
        employee_id = row[1].strip() or None
        phone = row[2].strip() if len(row) > 2 else ""
        designation = row[3].strip() if len(row) > 3 else ""
        process = row[4].strip().lower() if len(row) > 4 else ""
        
        raw_daily_rate = row[5].strip() if len(row) > 5 else ""
        raw_shift_hours = row[6].strip() if len(row) > 6 else ""
        salary_model = row[7].strip().upper() if len(row) > 7 else "DAILY"
        raw_fixed_sal = row[8].strip() if len(row) > 8 else ""
        raw_allowance = row[9].strip() if len(row) > 9 else ""
        raw_ot_rate = row[10].strip() if len(row) > 10 else ""
        
        identity_no = row[11].strip() if len(row) > 11 else ""
        emergency_name = row[12].strip() if len(row) > 12 else ""
        emergency_phone = row[13].strip() if len(row) > 13 else ""
        blood_group = row[14].strip() if len(row) > 14 else ""

        # 1. Check Required Fields
        if not name:
            errors.append("Worker Name is required.")
        if not process:
            errors.append("Process is required.")
        elif process not in ("casting", "machining", "polishing", "packaging"):
            errors.append("Process must be one of: casting, machining, polishing, packaging.")

        if salary_model not in ("DAILY", "FIXED"):
            errors.append("Salary Model must be DAILY or FIXED.")

        # 2. Check Deduplication
        if name:
            name_lower = name.lower()
            if name_lower in names_in_file:
                errors.append(f"Duplicate Worker Name '{name}' in upload file.")
            else:
                names_in_file.add(name_lower)

        if employee_id:
            if employee_id in emp_ids_in_file:
                errors.append(f"Duplicate Employee ID '{employee_id}' in upload file.")
            else:
                emp_ids_in_file.add(employee_id)

        # 3. Numeric conversions & bounds checking
        daily_rate = 0.0
        shift_hours = 8.0
        fixed_salary = 0.0
        allowance = 0.0
        ot_rate = 0.0

        try:
            daily_rate = to_float(raw_daily_rate, 0.0)
            if daily_rate < 0: errors.append("Daily rate cannot be negative.")
        except ValueError as e:
            errors.append(str(e))

        try:
            shift_hours = to_float(raw_shift_hours, 8.0)
            if shift_hours <= 0: errors.append("Standard shift hours must be greater than zero.")
        except ValueError as e:
            errors.append(str(e))

        try:
            fixed_salary = to_float(raw_fixed_sal, 0.0)
            if fixed_salary < 0: errors.append("Monthly fixed salary cannot be negative.")
        except ValueError as e:
            errors.append(str(e))

        try:
            allowance = to_float(raw_allowance, 0.0)
            if allowance < 0: errors.append("Monthly allowance cannot be negative.")
        except ValueError as e:
            errors.append(str(e))

        try:
            ot_rate = to_float(raw_ot_rate, 0.0)
            if ot_rate < 0: errors.append("Overtime rate cannot be negative.")
        except ValueError as e:
            errors.append(str(e))

        # 4. Parse Piece Rates
        rates = {}
        raw_rates = row[15].strip() if len(row) > 15 else ""
        if raw_rates:
            parsed_rates, parse_errors = parse_piece_rates(raw_rates)
            errors.extend(parse_errors)
            for code, rate_val in parsed_rates.items():
                if code not in item_codes:
                    errors.append(f"Item Code '{code}' not found in Item Master.")
                else:
                    rates[code] = rate_val

        # 5. Check if Worker already exists in DB to determine UPDATE vs INSERT
        action = "INSERT"
        if not errors:
            # Check by unique Employee ID if provided, otherwise fallback to exact Name
            if employee_id:
                exists = Worker.objects.filter(employee_id=employee_id).exists()
            else:
                exists = Worker.objects.filter(name=name).exists()
            if exists:
                action = "UPDATE"

        parsed_data = {
            "name": name,
            "employee_id": employee_id,
            "phone": phone,
            "designation": designation,
            "process": process,
            "daily_rate": daily_rate,
            "standard_shift_hours": shift_hours,
            "salary_model": salary_model,
            "monthly_fixed_salary": fixed_salary,
            "monthly_allowance": allowance,
            "overtime_rate": ot_rate,
            "identity_number": identity_no,
            "emergency_contact_name": emergency_name,
            "emergency_contact_phone": emergency_phone,
            "blood_group": blood_group,
            "rates": rates,
            "raw_rates": raw_rates
        }

        validated_rows.append({
            "row_idx": idx,
            "data": parsed_data,
            "action": "ERROR" if errors else action,
            "errors": errors
        })

    return validated_rows


def validate_job_workers_data(rows):
    """
    Validates rows of JobWorker data before import.
    Returns: list of dicts (parsed row + validation status)
    """
    from apps.workforce.models import JobWorker
    validated_rows = []
    codes_in_file = set()
    names_in_file = set()

    # Pre-fetch existing item codes from DB for validation
    item_codes = set(Item.objects.values_list('code', flat=True))

    for idx, row in enumerate(rows, 1):
        errors = []
        
        # Ensure row has enough columns
        if len(row) < 3:
            validated_rows.append({
                "row_idx": idx,
                "data": {"name": "N/A", "jw_code": "N/A"},
                "action": "ERROR",
                "errors": ["Row has too few columns."]
            })
            continue

        # Extract columns safely
        jw_code = row[0].strip()
        name = row[1].strip()
        phone = row[2].strip() if len(row) > 2 else ""
        email = row[3].strip() if len(row) > 3 else ""
        address = row[4].strip() if len(row) > 4 else ""
        gst_number = row[5].strip() if len(row) > 5 else ""
        process = row[6].strip().lower() if len(row) > 6 else ""
        raw_rates = row[7].strip() if len(row) > 7 else ""

        # 1. Check Required Fields
        if not name:
            errors.append("Name is required.")
        if not process:
            errors.append("Process is required.")
        elif process not in ("casting", "machining", "polishing", "packaging"):
            errors.append("Process must be one of: casting, machining, polishing, packaging.")

        # 2. Check Deduplication
        if jw_code:
            if jw_code in codes_in_file:
                errors.append(f"Duplicate Job Worker Code '{jw_code}' in upload file.")
            else:
                codes_in_file.add(jw_code)

        if name:
            name_lower = name.lower()
            if name_lower in names_in_file:
                errors.append(f"Duplicate Job Worker Name '{name}' in upload file.")
            else:
                names_in_file.add(name_lower)

        # 3. Parse Rates
        rates = {}
        if raw_rates:
            parsed_rates, parse_errors = parse_piece_rates(raw_rates)
            errors.extend(parse_errors)
            for code, rate_val in parsed_rates.items():
                if code not in item_codes:
                    errors.append(f"Item Code '{code}' not found in Item Master.")
                else:
                    rates[code] = rate_val

        # 4. Check if JobWorker already exists in DB to determine UPDATE vs INSERT
        action = "INSERT"
        if not errors:
            if jw_code:
                exists = JobWorker.objects.filter(jw_code=jw_code).exists()
            else:
                exists = JobWorker.objects.filter(name=name).exists()
            if exists:
                action = "UPDATE"

        parsed_data = {
            "jw_code": jw_code,
            "name": name,
            "phone": phone,
            "email": email,
            "address": address,
            "gst_number": gst_number,
            "process": process,
            "rates": rates,
            "raw_rates": raw_rates
        }

        validated_rows.append({
            "row_idx": idx,
            "data": parsed_data,
            "action": "ERROR" if errors else action,
            "errors": errors
        })

    return validated_rows


def commit_items_import(validated_rows):
    """
    Saves validated Items inside a transaction.
    """
    created_count = 0
    updated_count = 0

    with transaction.atomic():
        for row in validated_rows:
            if row["action"] == "ERROR":
                raise ValueError(f"Cannot commit import with unresolved validation errors on row {row['row_idx']}.")
            
            data = row["data"]
            item, created = Item.objects.update_or_create(
                code=data["code"],
                defaults={
                    "name": data["name"],
                    "category": data["category"],
                    "sub_category": data["sub_category"],
                    "material": data["material"],
                    "variant": data["variant"],
                    "casting_weight": data["casting_weight"],
                    "machining_weight": data["machining_weight"],
                    "rate_per_piece": data["rate_per_piece"],
                    "lot_size": data["lot_size"],
                    "lot_with_box": data["lot_with_box"],
                    "casting_required": data["casting_required"],
                    "machining_required": data["machining_required"],
                    "polishing_required": data["polishing_required"],
                    "packing_required": data["packing_required"],
                    "notes": data["notes"],
                    "client_id": data.get("client_id"),
                    "active": True
                }
            )
            
            # Set companies ManyToMany
            company_ids = data.get("company_ids")
            if company_ids:
                item.companies.set(company_ids)
            else:
                item.companies.clear()
                
            # Set process rates
            from apps.ledger_pay.models import ItemWorkerAllocation
            ItemWorkerAllocation.objects.filter(item=item).delete()
            
            # Casting Allocation
            casting_worker_id = data.get("casting_worker_id")
            casting_rate = data.get("casting_rate", 0.0)
            if casting_worker_id and casting_rate > 0:
                ItemWorkerAllocation.objects.create(
                    item=item,
                    worker_id=casting_worker_id,
                    rate_per_piece=casting_rate
                )
            
            # Machining Allocation
            machining_worker_id = data.get("machining_worker_id")
            machining_rate = data.get("machining_rate", 0.0)
            if machining_worker_id and machining_rate > 0:
                ItemWorkerAllocation.objects.create(
                    item=item,
                    worker_id=machining_worker_id,
                    rate_per_piece=machining_rate
                )
            
            # Polishing Allocation
            polishing_worker_id = data.get("polishing_worker_id")
            polishing_rate = data.get("polishing_rate", 0.0)
            if polishing_worker_id and polishing_rate > 0:
                ItemWorkerAllocation.objects.create(
                    item=item,
                    worker_id=polishing_worker_id,
                    rate_per_piece=polishing_rate
                )
            
            # Packing Allocation
            packing_worker_id = data.get("packing_worker_id")
            packing_rate = data.get("packing_rate", 0.0)
            if packing_worker_id and packing_rate > 0:
                ItemWorkerAllocation.objects.create(
                    item=item,
                    worker_id=packing_worker_id,
                    rate_per_piece=packing_rate
                )
            
            # Job Worker Allocation
            job_worker_id = data.get("job_worker_id")
            job_worker_rate = data.get("job_worker_rate", 0.0)
            if job_worker_id and job_worker_rate > 0:
                ItemWorkerAllocation.objects.create(
                    item=item,
                    job_worker_id=job_worker_id,
                    rate_per_piece=job_worker_rate
                )

            if created:
                created_count += 1
            else:
                updated_count += 1

    return created_count, updated_count


def commit_workers_import(validated_rows):
    """
    Saves validated Workers inside a transaction.
    """
    created_count = 0
    updated_count = 0

    with transaction.atomic():
        for row in validated_rows:
            if row["action"] == "ERROR":
                raise ValueError(f"Cannot commit import with unresolved validation errors on row {row['row_idx']}.")
            
            data = row["data"]
            # Locate unique record
            if data["employee_id"]:
                worker = Worker.objects.filter(employee_id=data["employee_id"]).first()
            else:
                worker = Worker.objects.filter(name=data["name"]).first()

            defaults = {
                "name": data["name"],
                "phone": data["phone"],
                "designation": data["designation"],
                "process": data["process"],
                "daily_rate": data["daily_rate"],
                "standard_shift_hours": data["standard_shift_hours"],
                "salary_model": data["salary_model"],
                "monthly_fixed_salary": data["monthly_fixed_salary"],
                "monthly_allowance": data["monthly_allowance"],
                "overtime_rate": data["overtime_rate"],
                "identity_number": data["identity_number"],
                "emergency_contact_name": data["emergency_contact_name"],
                "emergency_contact_phone": data["emergency_contact_phone"],
                "blood_group": data["blood_group"],
                "active": True
            }

            if data["employee_id"]:
                defaults["employee_id"] = data["employee_id"]

            if worker:
                for key, val in defaults.items():
                    setattr(worker, key, val)
                worker.save()
                updated_count += 1
            else:
                worker = Worker.objects.create(**defaults)
                created_count += 1

            # Save piece rates
            if "rates" in data and data["rates"]:
                from apps.ledger_pay.models import ItemWorkerAllocation
                ItemWorkerAllocation.objects.filter(worker=worker).delete()
                for code, rate in data["rates"].items():
                    item = Item.objects.get(code=code)
                    ItemWorkerAllocation.objects.create(
                        worker=worker,
                        item=item,
                        rate_per_piece=rate
                    )

    return created_count, updated_count


def commit_job_workers_import(validated_rows):
    """
    Saves validated Job Workers and their item rates inside a transaction.
    """
    from apps.workforce.models import JobWorker
    from apps.ledger_pay.models import ItemWorkerAllocation
    created_count = 0
    updated_count = 0

    with transaction.atomic():
        for row in validated_rows:
            if row["action"] == "ERROR":
                raise ValueError(f"Cannot commit import with unresolved validation errors on row {row['row_idx']}.")
            
            data = row["data"]
            jw_code = data["jw_code"]
            
            defaults = {
                "name": data["name"],
                "phone": data["phone"],
                "email": data["email"],
                "address": data["address"],
                "gst_number": data["gst_number"],
                "process": data["process"],
                "active": True
            }
            
            # Locate unique record
            if jw_code:
                jw = JobWorker.objects.filter(jw_code=jw_code).first()
            else:
                jw = JobWorker.objects.filter(name=data["name"]).first()

            if jw_code:
                defaults["jw_code"] = jw_code

            if jw:
                for key, val in defaults.items():
                    setattr(jw, key, val)
                jw.save()
                updated_count += 1
            else:
                jw = JobWorker.objects.create(**defaults)
                created_count += 1
                
            # Save rates
            if "rates" in data and data["rates"]:
                ItemWorkerAllocation.objects.filter(job_worker=jw).delete()
                for code, rate in data["rates"].items():
                    item = Item.objects.get(code=code)
                    ItemWorkerAllocation.objects.create(
                        job_worker=jw,
                        item=item,
                        rate_per_piece=rate
                    )

    return created_count, updated_count
